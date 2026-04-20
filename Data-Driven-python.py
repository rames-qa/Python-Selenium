from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import os

# Excel file path (FIXED)
path = r"C:\Users\rames\Favorites\eclipse-workspace\Selenium-java\src\test\resources\Data.xlsx"

# Check file exists
if not os.path.exists(path):
    print("File not found:", path)
    exit()

# Load Excel
workbook = load_workbook(path)
sheet = workbook.active

# Start browser
driver = webdriver.Chrome()

# Loop through Excel (skip header row)
for i in range(2, sheet.max_row + 1):

    username = sheet.cell(row=i, column=1).value
    password = sheet.cell(row=i, column=2).value

    # Handle empty cells
    username = "" if username is None else str(username)
    password = "" if password is None else str(password)

    # Open website
    driver.get("https://www.saucedemo.com/")

    # Enter data
    driver.find_element(By.ID, "user-name").clear()
    driver.find_element(By.ID, "user-name").send_keys(username)

    driver.find_element(By.ID, "password").clear()
    driver.find_element(By.ID, "password").send_keys(password)

    driver.find_element(By.ID, "login-button").click()

    # Wait to observe result
    time.sleep(2)

# Close browser
driver.quit()